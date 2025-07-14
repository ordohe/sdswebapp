from flask import Flask, render_template, request, send_file, flash, redirect, url_for, session
import os
import csv
import uuid
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from functools import wraps

app = Flask(__name__)
app.secret_key = 'supersecretkey'

app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'output'
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10 MB max

# Authentication settings
ACCESS_CODE = os.environ.get('ACCESS_CODE', 'sds2024')  # Set this in Render environment variables

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'authenticated' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        access_code = request.form.get('access_code')
        if access_code == ACCESS_CODE:
            session['authenticated'] = True
            flash('Successfully logged in!', 'success')
            return redirect(url_for('home'))
        else:
            flash('Invalid access code. Please try again.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('authenticated', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def home():
    return render_template('home.html')

@app.route('/split-by-artist')
@login_required
def split_by_artist():
    return render_template('upload.html')

@app.route('/pivot-table')
@login_required
def pivot_table():
    return render_template('pivottable.html')

@app.route('/inventory', methods=['GET', 'POST'])
@login_required
def inventory():
    inventory_file = os.path.join(app.config['UPLOAD_FOLDER'], 'slowdownsoundsstock.xlsx')
    data = []
    columns = []

    if request.method == 'POST':
        # Save the edited inventory
        try:
            # Get columns from the form
            columns = request.form.getlist('columns')
            # Get number of rows
            num_rows = int(request.form.get('num_rows', 0))
            # Build new data
            new_data = []
            for i in range(num_rows):
                row = []
                for col in columns:
                    row.append(request.form.get(f'cell_{i}_{col}', ''))
                new_data.append(row)
            # Write to Excel
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet()
            ws.append(columns)
            for row in new_data:
                ws.append(row)
            wb.save(inventory_file)
            wb.close()
            flash('Inventory updated successfully!')
        except Exception as e:
            flash(f'Error saving inventory: {str(e)}')

    # Always read the latest data
    if os.path.exists(inventory_file):
        try:
            wb = load_workbook(inventory_file)
            ws = wb.active
            if not ws:
                flash("No active worksheet found in inventory file.")
                return render_template('inventory.html', data=[], columns=[])
            for row in ws.iter_rows(values_only=True):
                if not columns:
                    columns = [str(cell) if cell else '' for cell in row]
                else:
                    row_data = {}
                    for i, cell in enumerate(row):
                        if i < len(columns):
                            row_data[columns[i]] = str(cell) if cell else ''
                    data.append(row_data)
            wb.close()
        except Exception as e:
            flash(f"Error reading inventory file: {str(e)}")
    return render_template('inventory.html', data=data, columns=columns)

@app.route('/upload-inventory', methods=['POST'])
@login_required
def upload_inventory():
    file = request.files.get('file')
    
    if not file:
        flash("No file uploaded.")
        return redirect(url_for('inventory'))
    
    # Always save as slowdownsoundsstock.xlsx
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'slowdownsoundsstock.xlsx')
    
    try:
        file.save(file_path)
        flash("Inventory file uploaded successfully!")
    except Exception as e:
        flash(f"Error uploading file: {str(e)}")
    
    return redirect(url_for('inventory'))

@app.route('/upload', methods=['POST'])  # type: ignore
@login_required
def upload_file():
    file = request.files.get('file')
    action = request.form.get('action')

    if not file or not action:
        flash("Missing file or action.")
        return redirect(url_for('split_by_artist'))

    filename = secure_filename(file.filename or 'uploaded_file')
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)
    base_name = os.path.splitext(filename)[0]

    try:
        if action == 'split':
            # Read data and find artist column
            data = []
            header_row_index = None
            artist_col_index = None
            
            if filename.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for i, row in enumerate(reader):
                        data.append(row)
                        if i < 10:  # Check first 10 rows for header
                            for j, cell in enumerate(row):
                                if 'artist' in str(cell).lower():
                                    header_row_index = i
                                    artist_col_index = j
                                    break
                        if header_row_index is not None:
                            break
            else:
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                if ws:
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        data.append(list(row))
                        if i < 10:  # Check first 10 rows for header
                            for j, cell in enumerate(row):
                                if 'artist' in str(cell).lower():
                                    header_row_index = i
                                    artist_col_index = j
                                    break
                            if header_row_index is not None:
                                break
                wb.close()

            if header_row_index is None or artist_col_index is None:
                flash("Could not find a row containing 'artist' in the first 10 rows.")
                return redirect(url_for('split_by_artist'))

            # Group data by artist
            artist_groups = {}
            for row in data[header_row_index + 1:]:
                if len(row) > artist_col_index and row[artist_col_index]:
                    artist = str(row[artist_col_index]).strip()
                    if artist:
                        if artist not in artist_groups:
                            artist_groups[artist] = []
                        artist_groups[artist].append(row)

            # Create Excel file with separate sheets for each artist
            excel_name = f"{secure_filename(base_name)}_split_by_artist.xlsx"
            excel_path = os.path.join(app.config['OUTPUT_FOLDER'], excel_name)
            
            wb = Workbook()
            if wb.active:
                wb.remove(wb.active)  # Remove default sheet
            
            for artist, rows in artist_groups.items():
                if not artist.strip():
                    continue
                safe_name = artist[:31].replace('/', '-').replace('\\', '-')
                ws = wb.create_sheet(title=safe_name)
                
                # Write header
                ws.append(data[header_row_index])
                # Write data
                for row in rows:
                    ws.append(row)
            
            wb.save(excel_path)
            return send_file(excel_path, as_attachment=True, download_name=excel_name)

        elif action == 'pivot':
            flash("Pivot table functionality temporarily disabled. Please use split by artist.")
            return redirect(url_for('pivot_table'))


    except Exception as e:
        flash(f"Error processing file: {str(e)}")
        if action == 'pivot':
            return redirect(url_for('pivot_table'))
        return redirect(url_for('split_by_artist'))

if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5001)))
